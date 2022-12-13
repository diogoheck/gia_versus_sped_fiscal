from openpyxl import Workbook, load_workbook
import pandas as pd


class PastaTrabalho:
    def __init__(self, nome_pasta):
        self.nome_pasta = nome_pasta + '.xlsx'

    def criar_nova_pasta_de_trabalho(self):
        wb = Workbook()
        wb.save(self.nome_pasta)


class Planilha:
    def __init__(self, nome_plan):
        self.nome_plan = nome_plan

    def __str__(self):
        return f'Criada nova planilha: {self.nome}'

    def criar_planilha_primeira(self, nome_pasta):
        wb = load_workbook(nome_pasta)
        ws = wb.active
        ws.title = self.nome_plan
        ws.append(['Data', 'CFOP', 'Valor', 'Tipo'])
        wb.save(nome_pasta)

    def criar_nova_planilha(self, nome_pasta):

        wb = load_workbook(nome_pasta)
        wb.create_sheet(self.nome_plan)
        ws = wb[self.nome_plan]
        ws.append(['Data', 'CFOP', 'Valor', 'Tipo'])
        wb.save(nome_pasta)

    def add_plan(self, lista, nome_pasta):
        wb = load_workbook(nome_pasta)
        ws = wb[self.nome_plan]

        for li in lista:
            ws.append(li)
        wb.save(nome_pasta)

    def agrupar_planilha(self, nome_pasta):
        df = pd.read_excel(
            nome_pasta, self.nome_plan, header=0)
        plan = df.groupby(['Data', 'CFOP', 'Tipo'])[
            ['Valor']].sum().reset_index()
        plan = plan.set_index('CFOP')
        # plan = plan.T
        writer = pd.ExcelWriter(nome_pasta, engine='xlsxwriter')
        plan.to_excel(writer, sheet_name=self.nome_plan)
        writer.save()
