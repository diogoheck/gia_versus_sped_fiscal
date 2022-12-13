from openpyxl import load_workbook
import pandas as pd
from openpyxl.worksheet.properties import WorksheetProperties, PageSetupProperties
from openpyxl.utils import get_column_letter

ARQUIVO_GERADO = 'C190.xlsx'


def autosize(ws):
    for i in range(1, ws.max_column+1):
        ws.column_dimensions[get_column_letter(i)].bestFit = True


def formata_numero(ws, intervalo):
    columns = ws[intervalo]

    for rows in columns:
        for cell in rows:
            cell.number_format = '#,##0.00'


def converte_em_valor(str_valor):
    str_valor = round(float(str_valor.replace(
        ',', '.')), 2) if str_valor else ''
    return str_valor


def adiciona_cabecalho_lista(lista):
    lista.append(['CNPJ', 'DATA', 'NUM_NOTA', 'CHAVE', 'CFOP', 'VALOR'])
    return lista


def salvar_excel(lista):
    wb = load_workbook(ARQUIVO_GERADO)
    ws = wb.active
    ws.title = 'C190'
    for li in lista:
        ws.append(li)

    formata_numero(ws, 'F1:F10')
    # autosize(ws)

    wb.save(ARQUIVO_GERADO)


def agrupar_planilha_pandas():
    nome_plan = 'C190'
    df = pd.read_excel(
        ARQUIVO_GERADO, sheet_name=nome_plan)
    plan = df.groupby(['CNPJ', 'DATA', 'NUM_NOTA', 'CHAVE', 'CFOP'])[
        ['VALOR']].sum().reset_index()
    writer = pd.ExcelWriter(ARQUIVO_GERADO, engine='xlsxwriter')
    # workbook = writer.book
    # worksheet = writer.sheets[df.Name]
    # number_format = workbook.add_format({'num_format': '#,##0.00'})
    # worksheet.set_column('F:F', 12, number_format)
    plan.to_excel(writer, sheet_name=nome_plan, index=False)

    writer.save()


def gravar_informacoes_excel(obj_sped):
    lista = []
    lista = adiciona_cabecalho_lista(lista)
    for nota in obj_sped.notas:
        IE = obj_sped.CNPJ
        data = nota.campos_c100[11]
        numero_nota = nota.campos_c100[8]
        chave_nota = nota.campos_c100[9]

        for iten in nota.itens_c190:

            CFOP = str(iten.campos_c190[3])
            valor = converte_em_valor(iten.campos_c190[5])
            # print(IE, data, numero_nota, chave_nota, CFOP, valor)
            lista.append([IE, data, numero_nota, chave_nota, CFOP, valor])
    salvar_excel(lista)
