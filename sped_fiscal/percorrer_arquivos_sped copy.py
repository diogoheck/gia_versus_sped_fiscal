import os
from subprocess import list2cmdline
import pandas as pd
from sped_fiscal import sped_fiscal as sp
from planilha_excel import criar_plan
from classes_preparacao_dados import objetos_operacoes_v_total_oper as dados
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Border, Side, PatternFill, Font, GradientFill, Alignment
from openpyxl.utils import get_column_letter
from time import sleep  

class Dics:
    dic_operacoes_saidas = {}
    dic_operacoes_entradas = {}


def eh_arquivo_txt(arquivo):
    if arquivo.split('.')[-1] == 'txt':
        return True
    return False


def converte_em_valor(str_valor):
    str_valor = round(float(str_valor.replace(
        ',', '.')), 2) if str_valor else ''
    return str_valor


def converte_em_data(str_date):
    if str_date:

        str_date = str_date[0:2] + '/' + str_date[2:4] + '/' + str_date[4:8]

    return str_date


def gera_datas(di, df):
    v_final = int(df[0:2])

    lista_datas = []
    for i in range(v_final):
        i = str(i + 1)
        if len(i) == 1:
            i = '0' + i
        lista_datas.append(i + di[2:4] + di[4:8])
    lista_datas.append('Total')

    return lista_datas


    #  gravar_informacoes_excel(
                            # novo_sped, dic, lista_tipos_inex, pasta_trabalho,
                            # data_sped_inicial, data_sped_final, nome_pasta_trabalho)

def gravar_informacoes_excel(obj_sped, dic,
                             lista_tipos_inex, pasta_trabalho):
    lista = []
    dic_operacoes_entradas = {}
    dic_operacoes_saidas = {}
    lista_datas = []
    
    print(obj_sped)
    sleep(5)
    with open('log.txt', 'a') as log:
        for nota in obj_sped.notas:

            data = nota.campos_c100[11]
            for iten in nota.itens_c190:

                CFOP = str(iten.campos_c190[3])
                valor = converte_em_valor(iten.campos_c190[5])
                operacao = dic.get(iten.campos_c190[3])

                CFOP_VERIF = int(CFOP[0])

                if CFOP_VERIF >= 5:

                    if operacao:
                        if not Dics.dic_operacoes_saidas.get(int(operacao.split('-')[0])):
                            Dics.dic_operacoes_saidas = dados.criar_nova_operacao(
                                Dics.dic_operacoes_saidas, operacao, CFOP, data)
                        if not Dics.dic_operacoes_saidas[int(operacao.split('-')[0])].get(CFOP):
                            Dics.dic_operacoes_saidas = dados.criar_novo_CFOP(
                                Dics.dic_operacoes_saidas, operacao, CFOP, data)
                        Dics.dic_operacoes_saidas = dados.atualizar_dados(
                            Dics.dic_operacoes_saidas, operacao, CFOP, data, valor)

                else:

                    if operacao:
                        if not Dics.dic_operacoes_entradas.get(int(operacao.split('-')[0])):
                            Dics.dic_operacoes_entradas = dados.criar_nova_operacao(
                                Dics.dic_operacoes_entradas, operacao, CFOP, data)
                        if not Dics.dic_operacoes_entradas[int(operacao.split('-')[0])].get(CFOP):
                            Dics.dic_operacoes_entradas = dados.criar_novo_CFOP(
                                Dics.dic_operacoes_entradas, operacao, CFOP, data)
                        Dics.dic_operacoes_entradas = dados.atualizar_dados(
                            Dics.dic_operacoes_entradas, operacao, CFOP, data, valor)

                if not operacao and (CFOP not in lista_tipos_inex):
                    lista_tipos_inex.append(CFOP)
                    print(f'CFOP {CFOP} nao cadastrado', file=log)
        if log:
            print('Arquivo gerado com sucesso!!', file=log)

        # dicionario gerado com erro
        # print(Dics.dic_operacoes_saidas)


        pasta_trabalho.inserir_dados(
            Dics.dic_operacoes_entradas, lista_tipos_inex)
        pasta_trabalho.criar_nova_plan(
            obj_sped.CNPJ, 0)
        pasta_trabalho.inserir_dados(
            Dics.dic_operacoes_saidas, lista_tipos_inex)
        pasta_trabalho.criar_nova_plan(
            obj_sped.CNPJ, 34)


def criar_novo_sped(razao_social, CNPJ, IE):
    novo_sped = sp.SpedFiscal(razao_social, CNPJ, IE)
    return novo_sped


def criar_nova_planilha():
    pass


def criar_nova_pasta_trabalho(nome_pasta):
    nova_pasta = criar_plan.PastaTrabalho(nome_pasta)
    nova_pasta.criar_nova_pasta_de_trabalho()
    return nova_pasta


def percorrer_arquivo_sped(pasta, dic, primeira_passada):

    lista_tipos_inex = []
    pasta_trabalho = criar_plan.PastaTrabalho(pd)

    for diretorio, subpastas, arquivos in os.walk(pasta):
        print(arquivos)
        for arquivo in arquivos:
            # pasta_trabalho = criar_plan.PastaTrabalho(pd)
            sped = os.path.join(diretorio, arquivo)
            if eh_arquivo_txt(arquivo.split('.')[-1]):
                with open(sped, encoding='ansi') as arquivo_sped:
                    # print(arquivo_sped)
                    for linha in arquivo_sped:
                        registro = linha.strip().split('|')
                        if registro[1] == '9999':
                            break
                        if registro[1] == '0000':
                            novo_sped = criar_novo_sped(
                                registro[6], registro[7], registro[10])
                            data_sped_inicial = registro[4]
                            data_sped_final = registro[5]
                            if primeira_passada:
                                pasta_trabalho.create_arquivo_excel(
                                    registro[6] + '.xlsx')
                                nome_pasta_trabalho = registro[6] + '.xlsx'
                                primeira_passada = False
                        elif registro[1] == 'C100':
                            # data_nota = registro[11]
                            numero_nota = registro[8]
                            novo_sped.add(registro)
                        elif registro[1] == 'C190':
                            nf = novo_sped.procurar(numero_nota)
                            nf.add_C190(registro)
                        
                gravar_informacoes_excel(
                            novo_sped, dic, lista_tipos_inex, pasta_trabalho)
    pasta_trabalho.salvar_excel()

    # ler o excel novamente para formatações
    my_fill = PatternFill(start_color='FFFF00',
                          end_color='FFFF00',
                          fill_type='solid')
    lista_oper = []
    for k, v in dic.items():
        lista_oper.append(v.split('-')[-1])
    pasta_trabalho = load_workbook(nome_pasta_trabalho)
    for planilhas in pasta_trabalho:
        for planilha in planilhas:
            for cell in planilha:
                if cell.value in lista_oper:
                    coluna_pintar = get_column_letter(cell.column)
                    linha_pintar = cell.row


def percorrer_todos_arquivos(pasta, dic):
    percorrer_arquivo_sped(pasta, dic, True)
